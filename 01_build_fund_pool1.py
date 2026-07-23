"""
01  基金池构建 —— 白名单维护版（Maintenance Mode）

数据源: 副本全市场量化指数基金.xlsx（由 01_build_fund_pool_full.py 全市场扫描初筛 + 人工复核精筛）
流程:
  1. 读取白名单 Excel（4 个 Sheet = 4 个基准指数）
  2. 逐只调用 AKShare API 富化信息（成立日期、规模、基金经理等）
  3. 交叉校验：API 基准 vs 白名单预期 → 输出校验报告
  4. 保存 fund_master.parquet

如需重新全市场扫描，请运行 01_build_fund_pool_full.py。
"""

import os
import time
import random
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
import pandas as pd
import akshare as ak

from config import (
    DATA_DIR,
    FUND_MASTER_PATH,
    INDEX_NAMES,
    BENCHMARK_INDEX_MAP,
)

# ── 白名单 Excel 路径 ──
# 优先读中文文件名（本地编辑），CI 环境用 ASCII 名
_WHITELIST_CN = os.path.join(DATA_DIR, "副本全市场量化指数基金.xlsx")
_WHITELIST_ASCII = os.path.join(DATA_DIR, "fund_whitelist.xlsx")
WHITELIST_EXCEL = _WHITELIST_CN if os.path.exists(_WHITELIST_CN) else _WHITELIST_ASCII

# Sheet 名 → 基准代码
SHEET_BM_MAP = {
    "沪深300": "HS300",
    "中证500": "ZZ500",
    "中证1000": "ZZ1000",
    "中证全指": "CSI_ALL",
}

CACHE_PATH = os.path.join(DATA_DIR, "fund_detail_cache.parquet")
VALIDATION_REPORT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output", "fund_verification_report.csv")

REQUEST_DELAY_MIN = 0.3
REQUEST_DELAY_MAX = 0.8
MAX_RETRIES = 2
API_TIMEOUT = 15  # 单次API调用超时秒数
CACHE_SAVE_INTERVAL = 50


# ═══════════════════════════════════════════════════════════
#  工具函数
# ═══════════════════════════════════════════════════════════

def ensure_dirs():
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(os.path.dirname(VALIDATION_REPORT), exist_ok=True)


def _random_sleep():
    time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))


def _find_col(df, *names):
    for c in df.columns:
        for n in names:
            if n in str(c):
                return c
    return None


# ═══════════════════════════════════════════════════════════
#  API 缓存
# ═══════════════════════════════════════════════════════════

def _load_cache():
    if not os.path.exists(CACHE_PATH):
        return {}
    try:
        df = pd.read_parquet(CACHE_PATH)
        cache = {}
        for _, row in df.iterrows():
            cache[row["fund_code"]] = row.to_dict()
        print(f"  已加载 API 缓存: {len(cache)} 条")
        return cache
    except Exception:
        return {}


def _save_cache(cache):
    if not cache:
        return
    try:
        df = pd.DataFrame(list(cache.values()))
        df.to_parquet(CACHE_PATH, index=False)
    except Exception as e:
        print(f"  缓存保存失败: {e}")


# ═══════════════════════════════════════════════════════════
#  读取白名单 Excel
# ═══════════════════════════════════════════════════════════

def load_whitelist() -> pd.DataFrame:
    """从 Excel 四个 Sheet 读取白名单，返回 DataFrame"""
    print("=" * 60)
    print("Step 1 / 3  读取白名单 Excel")
    print("=" * 60)

    if not os.path.exists(WHITELIST_EXCEL):
        raise FileNotFoundError(f"白名单文件不存在: {WHITELIST_EXCEL}")

    xls = pd.ExcelFile(WHITELIST_EXCEL)
    rows = []

    for sheet_name in xls.sheet_names:
        if sheet_name not in SHEET_BM_MAP:
            print(f"  [WARN] 未知 Sheet: {sheet_name}，跳过")
            continue
        bm = SHEET_BM_MAP[sheet_name]
        df = pd.read_excel(xls, sheet_name, dtype=str)
        # 兼容旧格式: 可能有 '代码' '简称' 列（带 .OF 后缀），也可能直接是基金代码
        code_col = next((c for c in df.columns if '代码' in str(c) or 'code' in str(c).lower()), df.columns[0])
        name_col = next((c for c in df.columns if '简称' in str(c) or 'name' in str(c).lower()), df.columns[1])

        for _, row in df.iterrows():
            code = str(row[code_col]).strip().replace(".OF", "").zfill(6)
            name = str(row[name_col]).strip()
            if code and name:
                rows.append({
                    "fund_code": code,
                    "fund_name": name,
                    "benchmark_index": bm,
                    "benchmark_name": INDEX_NAMES.get(bm, bm),
                })

    result = pd.DataFrame(rows).drop_duplicates(subset="fund_code", keep="first")
    print(f"  Sheet 数: {len(xls.sheet_names)}")
    for sheet_name in xls.sheet_names:
        if sheet_name in SHEET_BM_MAP:
            bm = SHEET_BM_MAP[sheet_name]
            cnt = len(result[result["benchmark_index"] == bm])
            print(f"    {sheet_name}: {cnt} 只")
    print(f"  合计: {len(result)} 只\n")
    return result


# ═══════════════════════════════════════════════════════════
#  API 富化 + 交叉校验
# ═══════════════════════════════════════════════════════════

def _fetch_fund_detail(code: str, cache: dict) -> dict | None:
    """调用 AKShare 获取基金详情（优先缓存，带超时）"""
    if code in cache:
        return cache[code]

    for attempt in range(MAX_RETRIES):
        try:
            if attempt > 0:
                time.sleep(1 + random.uniform(0, 0.5))
            else:
                _random_sleep()

            # 用线程超时防止 AKShare 调用卡死
            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(ak.fund_individual_basic_info_xq, symbol=code)
                df = future.result(timeout=API_TIMEOUT)

            if df is None or df.empty:
                raise ValueError("empty response")

            info = {}
            for _, row in df.iterrows():
                if len(row) >= 2:
                    info[str(row.iloc[0]).strip()] = str(row.iloc[1]).strip()

            detail = {
                "fund_code": code,
                "benchmark_raw": "",
                "objective_raw": "",
                "benchmark": "",
                "objective": "",
                "fund_company": "",
                "fund_manager": "",
                "fund_type": "",
                "launch_date": "",
                "scale": "",
            }

            for k, v in info.items():
                if v in ["", "-", "—", "暂无"]:
                    continue
                if "业绩比较基准" in k or "业绩基准" in k:
                    detail["benchmark_raw"] = v
                    detail["benchmark"] = v
                elif "投资目标" in k:
                    detail["objective_raw"] = v
                    detail["objective"] = v
                elif "基金公司" in k or "管理人" in k:
                    detail["fund_company"] = v
                elif "基金经理" in k:
                    detail["fund_manager"] = v
                elif "基金类型" in k:
                    detail["fund_type"] = v
                elif "成立" in k:
                    detail["launch_date"] = v
                elif "规模" in k or "资产净值" in k:
                    detail["scale"] = v

            cache[code] = detail
            return detail

        except FuturesTimeoutError:
            print(f"\n  [TIMEOUT] {code} API 超时 ({API_TIMEOUT}s)，跳过")
            continue
        except Exception:
            continue

    return None


def _verify_benchmark(detail: dict, expected_bm: str) -> tuple[bool, str]:
    """
    交叉校验：API 返回的业绩基准是否与白名单标注的基准一致。
    返回 (是否一致, 原因说明)
    """
    if detail is None:
        return False, "API 获取失败，无法校验"

    benchmark = detail.get("benchmark", "")
    fund_type = detail.get("fund_type", "")

    # 从业绩基准中匹配指数
    matched_bm = None
    for keyword, bm_code in BENCHMARK_INDEX_MAP.items():
        if keyword in benchmark:
            matched_bm = bm_code
            break

    if matched_bm is None:
        # 尝试从基金类型判断
        if "增强" not in fund_type and "指数" not in fund_type:
            return False, f"基金类型不符: {fund_type}"
        return True, "业绩基准未明确匹配，但类型为指数型，接受"

    if matched_bm == expected_bm:
        return True, f"基准一致: {INDEX_NAMES.get(matched_bm, matched_bm)}"

    # 特殊处理: 沪深300 vs 中证全指可能有交叉
    return False, f"基准不一致: API={INDEX_NAMES.get(matched_bm, matched_bm)}, 白名单={INDEX_NAMES.get(expected_bm, expected_bm)}"


def enrich_and_verify(fund_pool: pd.DataFrame) -> pd.DataFrame:
    """
    逐只调用 API 富化信息，并进行交叉校验。
    返回 enriched fund_master DataFrame。
    """
    print("=" * 60)
    print("Step 2 / 3  API 富化 + 交叉校验")
    print("=" * 60)

    cache = _load_cache()
    total = len(fund_pool)
    enriched = []
    warnings = []
    api_failed = []

    t_start = time.time()

    for i, (_, row) in enumerate(fund_pool.iterrows()):
        code = row["fund_code"]
        name = row["fund_name"]
        expected_bm = row["benchmark_index"]

        # 进度
        elapsed = time.time() - t_start
        if i > 0:
            eta = (elapsed / i) * (total - i)
            pct = (i + 1) / total * 100
            print(f"\r  [{i + 1:>4}/{total} {pct:>4.0f}%]  "
                  f"OK{len(enriched):>4}  FAIL{len(api_failed):>3}  "
                  f"ETA{eta:.0f}s  {code} {name[:20]}",
                  end="", flush=True)

        # 调用 API
        detail = _fetch_fund_detail(code, cache)

        entry = {
            "fund_code": code,
            "fund_name": name,
            "benchmark_index": expected_bm,
            "benchmark_name": INDEX_NAMES.get(expected_bm, expected_bm),
            "fund_company": detail.get("fund_company", "") if detail else "",
            "fund_manager": detail.get("fund_manager", "") if detail else "",
            "fund_type": detail.get("fund_type", "") if detail else "",
            "benchmark_raw": detail.get("benchmark_raw", "") if detail else "",
            "objective_raw": detail.get("objective_raw", "") if detail else "",
            "launch_date": detail.get("launch_date", "") if detail else "",
            "scale": detail.get("scale", "") if detail else "",
            "source": "whitelist_excel",
            "api_status": "success" if detail else "failed",
        }

        # 交叉校验
        passed, verify_reason = _verify_benchmark(detail, expected_bm)
        entry["verify_passed"] = passed
        entry["verify_reason"] = verify_reason

        if not passed:
            warnings.append(entry.copy())

        if detail is None:
            api_failed.append(code)
            entry["verify_reason"] = "API 获取失败"

        enriched.append(entry)

        # 定期存缓存
        if len(enriched) % CACHE_SAVE_INTERVAL == 0:
            _save_cache(cache)

    print()
    _save_cache(cache)

    elapsed = time.time() - t_start
    print(f"\n  耗时: {elapsed:.0f}s")
    print(f"  API 成功: {len(enriched) - len(api_failed)}")
    print(f"  API 失败: {len(api_failed)}")
    print(f"  基准校验未通过: {len(warnings)}")

    # 保存 API 失败记录
    if api_failed:
        api_fail_path = os.path.join(DATA_DIR, "api_failed.csv")
        pd.DataFrame({"fund_code": api_failed}).to_csv(api_fail_path, index=False, encoding="utf-8-sig")
        print(f"  API 失败记录: {api_fail_path} ({len(api_failed)} 只)")

    # 保存校验报告
    if warnings:
        df_warn = pd.DataFrame(warnings)
        os.makedirs(os.path.dirname(VALIDATION_REPORT), exist_ok=True)
        df_warn.to_csv(VALIDATION_REPORT, index=False, encoding="utf-8-sig")
        print(f"\n  校验报告: {VALIDATION_REPORT}")
        print(f"  --- 未通过列表 ---")
        for w in warnings[:20]:
            print(f"    {w['fund_code']} {w['fund_name'][:30]}  |  {w['verify_reason']}")
        if len(warnings) > 20:
            print(f"    ... 共 {len(warnings)} 条，详见报告")

    return pd.DataFrame(enriched)


# ═══════════════════════════════════════════════════════════
#  保存
# ═══════════════════════════════════════════════════════════

def save_fund_master(df: pd.DataFrame):
    print("\n" + "=" * 60)
    print("Step 3 / 3  保存基金主表")
    print("=" * 60)

    df = df.drop_duplicates(subset="fund_code", keep="first")

    # 排序
    idx_order = {"HS300": 0, "ZZ500": 1, "ZZ1000": 2, "CSI_ALL": 3}
    df["_sort"] = df["benchmark_index"].map(idx_order).fillna(9)
    df = df.sort_values(["_sort", "fund_code"]).drop(columns=["_sort"]).reset_index(drop=True)

    df.to_parquet(FUND_MASTER_PATH, index=False)
    print(f"  已保存: {FUND_MASTER_PATH}")
    print(f"  基金数量: {len(df)}")

    passed = df[df["verify_passed"] == True]
    failed = df[df["verify_passed"] == False]
    api_fail = df[df["api_status"] == "failed"]

    print(f"    校验通过: {len(passed)}")
    print(f"    校验未通过: {len(failed)}")
    print(f"    API 失败: {len(api_fail)}")

    for bm in ["HS300", "ZZ500", "ZZ1000", "CSI_ALL"]:
        cnt = len(df[df["benchmark_index"] == bm])
        print(f"    {INDEX_NAMES.get(bm, bm)}: {cnt} 只")

    if len(failed) > 0:
        print(f"\n  [NOTE] 校验未通过的基金仍保留在 fund_master 中，详见 {VALIDATION_REPORT}")


# ═══════════════════════════════════════════════════════════
#  回写 Excel（API 结果填充到表格）
# ═══════════════════════════════════════════════════════════

SHEET_BM_MAP_REV = {v: k for k, v in SHEET_BM_MAP.items()}


def _write_back_to_excel(df: pd.DataFrame):
    """把 launch_date / scale 写回 Excel 对应的 Sheet。API 失败的填'获取失败'。"""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    print("\n" + "=" * 60)
    print("Step 3.5  回写 Excel（填充成立时间 + 规模）")
    print("=" * 60)

    wb = load_workbook(WHITELIST_EXCEL)
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    filled_ok = 0
    filled_fail = 0

    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_BM_MAP:
            continue
        ws = wb[sheet_name]
        bm = SHEET_BM_MAP[sheet_name]

        # 确保第 3、4 列是成立时间和规模的表头
        ws.cell(row=1, column=3, value="上市日期" if ws.cell(row=1, column=3).value is None else ws.cell(row=1, column=3).value)
        ws.cell(row=1, column=4, value="规模" if ws.cell(row=1, column=4).value is None else ws.cell(row=1, column=4).value)

        for row_idx in range(2, ws.max_row + 1):
            code = str(ws.cell(row=row_idx, column=1).value or "").strip().replace(".OF", "").zfill(6)
            if not code:
                continue

            match = df[(df["fund_code"] == code) & (df["benchmark_index"] == bm)]
            if match.empty:
                continue

            row_data = match.iloc[0]
            launch = row_data.get("launch_date", "")
            scale = row_data.get("scale", "")
            api_ok = row_data.get("api_status", "") == "success"

            if api_ok and launch:
                ws.cell(row=row_idx, column=3).value = launch
                ws.cell(row=row_idx, column=4).value = scale if scale else ""
                filled_ok += 1
            else:
                ws.cell(row=row_idx, column=3).value = "获取失败"
                ws.cell(row=row_idx, column=4).value = "获取失败"
                for ci in [3, 4]:
                    ws.cell(row=row_idx, column=ci).fill = yellow_fill
                filled_fail += 1

    # 始终保存 ASCII 版本（CI 兼容），同步中文版（如有）
    wb.save(_WHITELIST_ASCII)
    if os.path.exists(_WHITELIST_CN):
        wb.save(_WHITELIST_CN)
    print(f"  填充成功: {filled_ok} 只")
    print(f"  获取失败（标黄）: {filled_fail} 只（请手动查询后填入，去除标黄）")
    print(f"  已保存: {_WHITELIST_ASCII}"
          + (f" + {os.path.basename(_WHITELIST_CN)}" if os.path.exists(_WHITELIST_CN) else ""))


# ═══════════════════════════════════════════════════════════
#  主入口
# ═══════════════════════════════════════════════════════════

def main():
    print("\n" + "█" * 60)
    print("█  指数增强基金池构建（白名单校验模式）")
    print("█  数据源: 副本全市场量化指数基金.xlsx")
    print("█  流程: 读取白名单 → API 富化 → 交叉校验 → 输出")
    print("█" * 60)

    ensure_dirs()

    # 1. 读取白名单
    fund_pool = load_whitelist()
    if fund_pool.empty:
        print("[ERROR] 白名单为空")
        return

    # 2. API 富化 + 校验
    fund_master = enrich_and_verify(fund_pool)

    # 3. 保存
    save_fund_master(fund_master)

    # 3.5 回写 Excel
    _write_back_to_excel(fund_master)

    print("\n" + "█" * 60)
    print(f"[OK] 基金池构建完成: {len(fund_master)} 只")
    print(f"[OUTPUT] {FUND_MASTER_PATH}")
    if os.path.exists(VALIDATION_REPORT):
        print(f"[REPORT] {VALIDATION_REPORT}")
    print("█" * 60)
    print("\n下一步: 运行 02_update_nav.py 更新净值数据")


if __name__ == "__main__":
    main()
